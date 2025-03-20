import os.path
from datetime import datetime

import pandas as pd
from openpyxl.styles import Font, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from ttkbootstrap.dialogs import Messagebox

from file_manager import FileManager


class FileMerger():

    @staticmethod
    def read_files(first_file_name, second_file_name):
        base_path = FileManager.get_base_path_merge()
        first_file_path = os.path.join(base_path, first_file_name)
        second_file_path = os.path.join(base_path, second_file_name)

        first_df = FileManager.load_excel(first_file_path)
        second_df = FileManager.load_excel(second_file_path)

        if first_df is None or second_df is None:
            Messagebox.show_error("Jeden z wybranych plików nie posiada zawartości", "Błąd")
            return

        FileMerger.proceed_df(first_df, second_df, first_file_name)

    @staticmethod
    def proceed_df(first_df, second_df, first_file_name):

        merged_df = pd.concat([second_df, first_df], ignore_index=True)

        merged_df = FileMerger.replace_data(merged_df)

        report_df = FileMerger.transform_and_group_orders(merged_df)

        saved_file = FileManager.save_excel_file(
            report_df,
            FileManager.get_base_path_kw(),
            first_file_name[:5] + " LL.xlsx"
        )

        FileMerger.format_file(saved_file)

    @staticmethod
    def replace_data(merged_df):
        column_name = merged_df.columns[9]
        merged_df[column_name] = merged_df[column_name].str.replace("KOD", "KON")
        merged_df[column_name] = merged_df[column_name].str.replace("GYD", "GYA")

        return merged_df

    @staticmethod
    def transform_and_group_orders(df):
        country_map = {
            "SK": "01",
            "SI": "02",
            "CZ": "03",
            "HR": "04",
            "RO": "05",
            "BG": "06",
            "LT": "07",
            "LV": "07",
            "RS": "08",
            "PL": "09",
            "NI": "10",
        }

        result_rows = []

        for order, group in df.groupby("Rendelési szám"):
            main_location = group[group["Infó"] == "HAVI"]
            main_quantity = main_location["Raklapok \nszáma"].sum() if not main_location.empty else 0

            additional_locations = group[group["Infó"] != "HAVI"]

            if not additional_locations.empty:
                grouped_additional = (
                    additional_locations.groupby("Infó", as_index=False)["Raklapok \nszáma"]
                    .sum()
                )

                additional_text = " ".join(
                    f"{row['Infó']}({row['Raklapok \nszáma']}pal)"
                    for _, row in grouped_additional.iterrows()
                )
            else:
                additional_text = ""

            group["B"] = group["A címzett \nraktára"]
            group["A"] = group["A fogadó \nországa"].map(country_map).astype(str) + " " + group["A fogadó \nországa"]
            group["C"] = group["Indítás\n dátuma"]

            result_rows.append([
                group["A"].iloc[0],
                group["B"].iloc[0],
                group["C"].iloc[0] + str(datetime.now().year),
                order[:2],
                "",
                "",
                "",
                main_quantity,
                additional_text,
                order,
                "",
                "",
            ])

        result_df = pd.DataFrame(result_rows, columns=[
            "A fogadó ", "A címzett", "Indítás", "Rendelési \nszám", "Marcos no", "Firma", "reg no", "Gyal", "Comment",
            "Order no", "1", "2",
        ])

        return result_df

    @staticmethod
    def format_file(saved_file):
        wb = FileManager.open_excel_file(saved_file)
        ws = wb.active
        ws_new = wb.create_sheet('LoadingList')

        headers = [cell.value for cell in ws[1]]

        data = [list(row) for row in ws.iter_rows(min_row=2, values_only=True)]
        try:
            data.sort(key=lambda x: (x[2]))
            data.sort(key=lambda x: (x[3]))
            data.sort(key=lambda x: (x[1]))
            data.sort(key=lambda x: (x[0]))
        except Exception as e:
            print(f"Błąd podczas sortowania: {e}")

        font_colors = {
            "TL": "0000FF",  # Niebieski
            "KL": "FF0000",  # Czerwony
            "KT": "008000",  # Zielony
        }
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                             top=Side(style='thin'), bottom=Side(style='thin'))

        col_max_length = [len(header) for header in headers]
        ws_new.append(headers)

        for cell in ws_new[1]:
            cell.font = Font(name="Arial", size = 10, bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border

        for row_data in data:
            ws_new.append(row_data)
            row = ws_new[ws_new.max_row]

            order_value = row[3].value
            font_color = font_colors.get(order_value, "000000")

            for idx, cell in enumerate(row):
                cell.font = Font(name="Arial", size=10, color=font_color)
                cell.border = thin_border

                try:
                    col_max_length[idx] = max(col_max_length[idx], len(str(cell.value)))
                except:
                    pass

        for col_idx, max_length in enumerate(col_max_length, start=1):
            col_letter = get_column_letter(col_idx)
            ws_new.column_dimensions[col_letter].width = max_length + 2

        wb.remove(ws)
        wb.save(saved_file)
        Messagebox.show_info("Zapisano", f"Plik {saved_file} sformatowany i zapisany.")